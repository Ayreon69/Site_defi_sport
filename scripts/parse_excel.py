from __future__ import annotations

import csv
import glob
import json
import math
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl

MONTH_MAP = {
    "juin": 6,
    "aout": 8,
    "août": 8,
    "octobre": 10,
    "debut janvier": 1,
    "début janvier": 1,
}
START_YEAR = 2025

HEADER_MAP = {
    "dips": "dips",
    "pompes": "pompes",
    "traction pronation": "traction_pro",
    "traction supination": "traction_sup",
    "planche (min)": "planche_sec",
    "superman (min)": "superman_sec",
    "100 m (sec)": "sprint_100m_sec",
    "5 km (min)": "run_5km_sec",
    "poids (kg)": "poids",
}

OUTPUT_COLUMNS = [
    "personne",
    "date",
    "type",
    "dips",
    "pompes",
    "traction_pro",
    "traction_sup",
    "planche_sec",
    "superman_sec",
    "sprint_100m_sec",
    "run_5km_sec",
    "poids",
]


@dataclass
class CleanRow:
    personne: str
    date: str
    type: str
    dips: float | None
    pompes: float | None
    traction_pro: float | None
    traction_sup: float | None
    planche_sec: float | None
    superman_sec: float | None
    sprint_100m_sec: float | None
    run_5km_sec: float | None
    poids: float | None


@dataclass
class PersonMeta:
    personne: str
    gage: str | None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_month_to_date(value: Any, start_year: int = START_YEAR) -> str | None:
    month_label = normalize_text(value)
    if month_label not in MONTH_MAP:
        return None

    month = MONTH_MAP[month_label]
    first_month = 6
    year = start_year if month >= first_month else start_year + 1
    return date(year, month, 1).isoformat()


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    raw = str(value).strip()
    if not raw or raw.lower() in {"x", "na", "n/a", "none", "null", "-"}:
        return None

    raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def parse_duration_to_seconds(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, timedelta):
        return round(value.total_seconds(), 3)

    if isinstance(value, datetime):
        return round(value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000, 3)

    if isinstance(value, time):
        return round(value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000, 3)

    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    if not raw or raw.lower() in {"x", "na", "n/a", "none", "null", "-"}:
        return None

    if re.fullmatch(r"\d+(?:\.\d+)?", raw):
        return float(raw)

    parts = raw.split(":")
    if len(parts) in {2, 3}:
        try:
            parts_f = [float(p) for p in parts]
        except ValueError:
            return None

        if len(parts_f) == 2:
            minutes, seconds = parts_f
            return round(minutes * 60 + seconds, 3)

        hours, minutes, seconds = parts_f
        return round(hours * 3600 + minutes * 60 + seconds, 3)

    return None


def extract_block_rows(ws: Any, section_label: str, row_start_offset: int = 3, row_count: int = 4) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    target = normalize_text(section_label)

    for row in range(1, ws.max_row + 1):
        marker = normalize_text(ws.cell(row=row, column=2).value)
        if marker == target:
            first_data_row = row + row_start_offset
            for i in range(row_count):
                r = first_data_row + i
                month_text = ws.cell(row=r, column=2).value
                month_date = parse_month_to_date(month_text)
                if month_date:
                    rows.append((r, month_date))
            break

    return rows


def map_headers(ws: Any, header_row: int) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for col in range(3, 12):
        header = normalize_text(ws.cell(row=header_row, column=col).value)
        if header in HEADER_MAP:
            mapping[col] = HEADER_MAP[header]
    return mapping


def extract_sheet_rows(ws: Any, person_name: str) -> list[CleanRow]:
    output: list[CleanRow] = []

    section_configs = [
        ("RÉALISATION", "realisation"),
        ("PRÉVISIONNEL", "previsionnel"),
    ]

    for section_label, section_type in section_configs:
        block_rows = extract_block_rows(ws, section_label)
        if not block_rows:
            continue

        header_row = block_rows[0][0] - 1
        headers = map_headers(ws, header_row)

        for sheet_row, iso_date in block_rows:
            values: dict[str, float | None] = {k: None for k in OUTPUT_COLUMNS if k not in {"personne", "date", "type"}}
            for col, target_field in headers.items():
                cell_value = ws.cell(row=sheet_row, column=col).value
                if target_field in {"planche_sec", "superman_sec", "run_5km_sec"}:
                    values[target_field] = parse_duration_to_seconds(cell_value)
                elif target_field == "sprint_100m_sec":
                    values[target_field] = parse_float(cell_value)
                else:
                    values[target_field] = parse_float(cell_value)

            output.append(
                CleanRow(
                    personne=person_name,
                    date=iso_date,
                    type=section_type,
                    dips=values["dips"],
                    pompes=values["pompes"],
                    traction_pro=values["traction_pro"],
                    traction_sup=values["traction_sup"],
                    planche_sec=values["planche_sec"],
                    superman_sec=values["superman_sec"],
                    sprint_100m_sec=values["sprint_100m_sec"],
                    run_5km_sec=values["run_5km_sec"],
                    poids=values["poids"],
                )
            )

    return output


def extract_gage(ws: Any) -> str | None:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            marker = normalize_text(ws.cell(row=row, column=col).value)
            if marker == "gage":
                raw = ws.cell(row=row, column=col + 1).value
                if raw is None:
                    return None
                value = str(raw).strip()
                return value if value else None
    return None


def write_outputs(rows: list[CleanRow], meta: list[PersonMeta], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "clean_data.json"
    csv_path = out_dir / "clean_data.csv"
    meta_path = out_dir / "people_meta.json"

    payload = [asdict(row) for row in rows]
    meta_payload = [asdict(item) for item in meta]

    with json_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)

    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in payload:
            writer.writerow(row)

    with meta_path.open("w", encoding="utf-8") as fh:
        json.dump(meta_payload, fh, ensure_ascii=False, indent=2)


def main() -> None:
    candidates = glob.glob("*.xlsx")
    if not candidates:
        raise FileNotFoundError("Aucun fichier .xlsx trouvé à la racine du projet.")

    workbook_path = candidates[0]
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    all_rows: list[CleanRow] = []
    people_meta: list[PersonMeta] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows.extend(extract_sheet_rows(ws, sheet_name))
        people_meta.append(PersonMeta(personne=sheet_name, gage=extract_gage(ws)))

    all_rows.sort(key=lambda r: (r.personne.lower(), r.type, r.date))
    people_meta.sort(key=lambda p: p.personne.lower())
    write_outputs(all_rows, people_meta, Path("data"))

    print(f"Fichier source: {workbook_path}")
    print(f"Lignes générées: {len(all_rows)}")
    print("Sortie: data/clean_data.csv + data/clean_data.json + data/people_meta.json")


if __name__ == "__main__":
    main()
